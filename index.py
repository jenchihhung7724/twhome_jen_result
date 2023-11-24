import openpyxl 
import mysql.connector
import json


mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    password="1234",
    database="test"
)



def print_db():
    mycursor = mydb.cursor()

    mycursor.execute("SELECT * FROM jen_intergrated_1002")

    myresult = mycursor.fetchall()

    for x in myresult:
        
        data = []
        lists = []
        
        community_name = x[5] if x[5] else x[4]
        buildings_scrapy_ids = x[10] if x[10] else ""
        buildings_temp_id = x[11] if x[11] else 0

        community_names_from_i = get_communities_from_integreted_table(buildings_scrapy_ids)
        community_names_from_temp = get_communities_from_temp(buildings_temp_id)
        communities_from_itself = get_communities_from_itself(x[1])
       
        lists.append(x[4])
        if(x[5]):lists.append(x[5])
        if(community_names_from_i):lists.extend(community_names_from_i)
        if(community_names_from_temp):lists.extend(community_names_from_temp)
        if(communities_from_itself):lists.extend(communities_from_itself)
        lists = list(set(lists))
        data.append(x[1])
        data.append(x[15])
        data.append(x[6])
        data.append(community_name)
        data.append(lists)
        # print(data)
        # print(x[1] + x[15] + x[6] + x[4] + community_name + buildings_scrapy_ids)

def get_scrapy_ids(id):
    mycursor = mydb.cursor()

    mycursor.execute("SELECT buildings_scrapy_id FROM jen_intergrated_1002 where id = " + str(id))

    myresult = mycursor.fetchall()

    r = myresult[0][0] if myresult[0][0] else "[]"
    result = json.loads(r)
    return result

def get_related_scraoy_ids(sid):
    result = []
    mycursor = mydb.cursor()

    mycursor.execute("SELECT community_name, buildings_scrapy_id, buildings_temp_id,source_id FROM jen_intergrated_1002 where to_remove = '是' AND remove_reason LIKE '%" + str(sid) + "%'")


    myresult = mycursor.fetchall()
    for r in myresult:
        re = r[1] if r[1] else '[]'
        result.extend(json.loads(re))

    return result

def get_all_integrated_ids(id,sid):
    ids = []
    ids.extend(get_scrapy_ids(id))
    ids.extend(get_related_scraoy_ids(sid))
    return ids

def get_all_integrated_ids_as_json(ids):
    result = dict()

    for id in ids:
        mycursor = mydb.cursor()

        mycursor.execute("SELECT community_name FROM buildings_scrapy where id = " + str(id))

        myresult = mycursor.fetchall()
        result[id] = myresult[0][0]

    return result

def get_integrated_json(id,sid):
    ids = get_all_integrated_ids(id,sid)
    result = get_all_integrated_ids_as_json(ids)
    

    return json.dumps(result,ensure_ascii=False)

def get_communities_from_integreted_table(ids):
    new_ids = ids.replace("\"", "")
    new_ids = new_ids.replace("[","(")
    new_ids = new_ids.replace("]",")")
    if not new_ids:return    
    mycursor = mydb.cursor()

    mycursor.execute("SELECT community_name, transfer_id FROM buildings_scrapy where id in" + new_ids)

    myresult = mycursor.fetchall()
    output = {}
    output['communities'] = []
    output['community_id_591'] = 0

    for result in myresult:
        output['communities'].append(result[0])

    return output


def get_communities_from_temp(id):
    if not id:return    
    mycursor = mydb.cursor()
    mycursor.execute("SELECT community_name FROM buildings_temp where id = " + id)

    myresult = mycursor.fetchall()
    output = []
    for result in myresult:
        output.append(result[0])
    return output

def get_communities_from_itself(sid):
    if not sid:return    
    mycursor = mydb.cursor()
    mycursor.execute("SELECT community_name, buildings_scrapy_id, buildings_temp_id,source_id FROM jen_intergrated_1002 where to_remove = '是' AND remove_reason LIKE '%" + str(sid) + "%'")

    myresult = mycursor.fetchall()
    buildings_scrapy_ids = myresult[0][1] if myresult else ""
    buildings_temp_id = myresult[0][2] if myresult else 0

    communities_from_integreted_table = []
    if(buildings_scrapy_ids):
        communities_from_integreted_table_object = get_communities_from_integreted_table(buildings_scrapy_ids)
        if(communities_from_integreted_table_object):
            communities_from_integreted_table = communities_from_integreted_table_object['communities']

    communities_from_temp = []
    if(buildings_temp_id):
        communities_from_temp = get_communities_from_temp(buildings_temp_id)

    output = []
    for result in myresult:
        output.append(result[0])

    output.extend(communities_from_integreted_table)
    output.extend(communities_from_temp)

    return output

# print_db()



def stap2():   
    val = []
    mycursor = mydb.cursor()

    mycursor.execute("SELECT * FROM jen_intergrated_1002")

    myresult = mycursor.fetchall()

    for x in myresult:
        if x[3]:
            continue
        data = []
        lists = []
        community_name = x[5] if x[5] else x[4]
        buildings_scrapy_ids = x[10] if x[10] else ""
        buildings_temp_id = x[11] if x[11] else 0

        community_names_from_i_object = get_communities_from_integreted_table(buildings_scrapy_ids)
        if(community_names_from_i_object):
            community_names_from_i = community_names_from_i_object['communities'] 
        community_names_from_temp = get_communities_from_temp(buildings_temp_id)
        communities_from_itself = get_communities_from_itself(x[1])
      
        lists.append(x[4])
        if(x[5]):lists.append(x[5])
        if(community_names_from_i):lists.extend(community_names_from_i)
        if(community_names_from_temp):lists.extend(community_names_from_temp)
        if(communities_from_itself):lists.extend(communities_from_itself)
        lists = json.dumps(list(set(lists)),ensure_ascii=False)
        data.append(x[1])
        data.append(x[15])
        data.append(x[6])
        data.append(community_name)
        data.append(lists)

        integrated_json = get_integrated_json(x[0],x[1])
        data.append(integrated_json)
        data.append(buildings_temp_id)

        val.append(data)

    mycursor = mydb.cursor()
    sql = "INSERT INTO jen_integrated_result (buildings_scrapy_integrate_id, city, district,community_name,lists,list_sources_from_scrapy,building_temp_community_id) VALUES (%s, %s,%s, %s,%s,%s,%s)"
    mycursor.executemany(sql, val)

    mydb.commit()



def stap1():
    paths = [ "python_tainan.xlsx"]
  
    val = []

    for path in paths:
        wb_obj = openpyxl.load_workbook(path) 
        for sheetname in wb_obj.sheetnames:
            sheet_obj = wb_obj[sheetname]   
            row = sheet_obj.max_row
            column = sheet_obj.max_column
            for i in range(1, row + 1): 
                id = sheet_obj.cell(row = i, column = 1).value
                if(id == 'ID' or id == 66991):
                    continue

                data = []
                for j in range(1, column + 1): 
                    cell_obj = sheet_obj.cell(row = i, column = j) 
                    data.append(cell_obj.value)

                data.append(sheetname)
                print(data)
                val.append(tuple(data))

    mycursor = mydb.cursor()
    sql = "INSERT INTO jen_intergrated_1002 (source_id, group_id, to_remove,community_name, modified_community_name, district,address, buildings_scrapy_integrate_built_date,lat, lng,buildings_scrapy_id, buildings_temp_id, buildings_temp_built_date, remove_reason,city) VALUES (%s, %s,%s, %s,%s,%s, %s,%s, %s,%s, %s,%s, %s, %s, %s)"
    mycursor.executemany(sql, val)

    mydb.commit()

stap2()

def test():
    mycursor = mydb.cursor()

    mycursor.execute("SELECT community_name FROM buildings_scrapy where id in (58121,171638)")

    myresult = mycursor.fetchall()

    for x in myresult:
         print(x)






